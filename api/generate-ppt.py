"""Vercel Python serverless function — HCT-COHS KPI PPT Generator.
Fetches live data from Smartsheet API and generates downloadable .pptx files.
"""

import os, re, io, json, zipfile, tempfile
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from urllib.request import Request, urlopen
from datetime import datetime
from xml.etree import ElementTree as ET

# ── Namespaces ──
NS = {
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
    'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
}
for prefix, uri in NS.items():
    ET.register_namespace(prefix, uri)
ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
ET.register_namespace('c14', 'http://schemas.microsoft.com/office/drawing/2007/8/2/chart')
ET.register_namespace('c15', 'http://schemas.microsoft.com/office/drawing/2012/chart')
ET.register_namespace('c16r2', 'http://schemas.microsoft.com/office/drawing/2015/06/chart')

# ── Regions ──
REGIONS = {
    'AD Al Ain':      {'sheets': ['AAF','AAZ'], 'short': ['Falaj Hazza','Zakhir'],       'subtitle': 'Al Ain Falaj Hazza & Al Ain Zakhir'},
    'Abu Dhabi':      {'sheets': ['ADA','ADB'], 'short': ['Baniyas A','Baniyas B'],      'subtitle': 'Abu Dhabi Baniyas A & Abu Dhabi Baniyas B'},
    'AD Remote':      {'sheets': ['ADH','MZY'], 'short': ['Al Dhanna','Madinat Zayed'],  'subtitle': 'Al Dhanna Ruwais & Al Dhafra Madinat Zayed City'},
    'Dubai':          {'sheets': ['DMC','DBN'], 'short': ['Academic City','Al Nahda'],    'subtitle': 'Dubai Academic City & Dubai Al Nahda'},
    'Fujairah':       {'sheets': ['FJF','FJH'], 'short': ['Faseel','Hulaifat'],           'subtitle': 'Fujairah Faseel & Fujairah Hulaifat'},
    'Sharjah':        {'sheets': ['SJA','SJB'], 'short': ['Campus A','Campus B'],         'subtitle': 'Sharjah Campus A & Sharjah Campus B'},
    'Ras Al Khaimah': {'sheets': ['RKA','RKB'], 'short': ['Campus A','Campus B'],         'subtitle': 'RAK Campus A & RAK Campus B'},
}

# ── KPI structure ──
PILLAR_KPIS = [
    {'pillar': 'Leadership', 'weight': 0.20, 'rows': [2,3,4,5,6]},
    {'pillar': 'Risk Mgmt',  'weight': 0.20, 'rows': [7,8,9]},
    {'pillar': 'Training',   'weight': 0.10, 'rows': [10,11]},
    {'pillar': 'OCP & Emerg','weight': 0.25, 'rows': [12,13,14,15]},
    {'pillar': 'Perf Eval',  'weight': 0.25, 'rows': [16,17,18,19]},
]

KPI_WEIGHTS = {
    2: 0.30, 3: 0.10, 4: 0.10, 5: 0.25, 6: 0.25,
    7: 0.30, 8: 0.50, 9: 0.20,
    10: 0.50, 11: 0.50,
    12: 0.40, 13: 0.40, 14: 0.10, 15: 0.10,
    16: 0.30, 17: 0.30, 18: 0.20, 19: 0.20,
}

CHART_KPI_MAP = {
    'chart5.xml': 0,  'chart6.xml': 2,
    'chart8.xml': 3,  'chart7.xml': 4,
    'chart11.xml': 5, 'chart12.xml': 6, 'chart13.xml': 7,
    'chart16.xml': 8,
    'chart20.xml': 10, 'chart21.xml': 11,
    'chart22.xml': 12, 'chart23.xml': 13,
    'chart27.xml': 14, 'chart28.xml': 15,
    'chart29.xml': 17, 'chart30.xml': 16,
}

SINGLE_SERIES_CHARTS = {'chart7.xml', 'chart8.xml'}
MERGE_CAMPUS_CHARTS = {'chart7.xml', 'chart8.xml'}
REMOVE_UNDERLINE_CHARTS = {'chart5.xml'}

PIE_CHARTS = {'chart1.xml','chart2.xml','chart3.xml','chart4.xml',
              'chart9.xml','chart10.xml','chart14.xml','chart15.xml',
              'chart17.xml','chart18.xml','chart19.xml','chart24.xml',
              'chart25.xml','chart26.xml'}

# ── Smartsheet source → KPI row mapping ──
SYNC_SOURCES = [
    {'key': 'v2_hs_kpi_report', 'reportId': '4811266391494532', 'campusCol': 'Campuses', 'monthCol': 'Primary', 'valueCol': 'Submitted', 'kpi_row': 2},
    {'key': 'v2_external_compliance', 'sheetId': '4198632256393092', 'campusCol': 'Campus Code', 'monthCol': 'Primary', 'plannedCol': 'Applicable Compliance', 'actualCol': 'Actual Compliance', 'kpi_row': 4},
    {'key': 'v2_hs_committee', 'sheetId': '435993944477572', 'campusCol': 'Committee', 'monthCol': 'Reporting Month', 'plannedCol': 'Meeting Planned', 'actualCol': 'Meeting Conducted', 'kpi_row': 5},
    {'key': 'v2_hazard_id', 'sheetId': '7323092115214212', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month', 'plannedCol': 'Total Controls Identified', 'actualCol': 'Implemented Controls', 'kpi_row': 7},
    {'key': 'v2_risk_closed', 'sheetId': '7323092115214212', 'campusCol': 'Campus', 'monthCol': 'Primary', 'plannedCol': 'Total Risk Assessments Registered', 'actualCol': 'Risk Assessment Closed', 'kpi_row': 8},
    {'key': 'v2_risk_validated', 'sheetId': '7323092115214212', 'campusCol': 'Campus', 'monthCol': 'Primary', 'plannedCol': 'Total Assessments Register', 'actualCol': 'RA Validated and Signed Off', 'kpi_row': 9},
    {'key': 'v2_planned_training', 'reportId': '5332685084905348', 'campusCol': 'Campus', 'monthCol': 'Primary', 'plannedCol': 'Planned Training', 'actualCol': 'Training Conducted', 'kpi_row': 10},
    {'key': 'v2_safe_working', 'sheetId': '1693592581001092', 'campusCol': 'Campus', 'monthCol': 'Primary', 'plannedCol': 'No. of SOPs Verified', 'actualCol': 'No. of SOPs Implemented', 'kpi_row': 12},
    {'key': 'v2_drills', 'sheetId': '5053158949605252', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month', 'plannedCol': 'Planned Drill? (Yes/No)', 'actualCol': 'Are there any submission?', 'kpi_row': 13, 'yesNoCount': True},
    {'key': 'v2_permit_to_work', 'sheetId': '5899016251330436', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month', 'plannedCol': 'No. of PTWs Issued', 'actualCol': 'Total Work Registered', 'kpi_row': 14},
    {'key': 'v2_onsite_induction', 'sheetId': '5899016251330436', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month', 'plannedCol': "No. of New Contractors (Individuals)", 'actualCol': 'Contractors Inducted in the Reporting Month', 'kpi_row': 15},
    {'key': 'v2_ehs_inspection', 'sheetId': '4947401822392196', 'campusCol': 'Campus Code', 'monthCol': 'Primary', 'plannedCol': 'No. of EHS Inspections Planned', 'actualCol': 'No. of EHS Inspections Completed', 'kpi_row': 16},
    {'key': 'v2_findings_on_time', 'sheetId': '4947401822392196', 'campusCol': 'Campus Code', 'monthCol': 'Primary', 'plannedCol': 'No. of Findings in Reporting Month', 'actualCol': 'No. of Findings Due', 'kpi_row': 17},
    {'key': 'v2_investigation_on_time', 'reportId': '6831846506581892', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month', 'plannedCol': 'Total Incident Investigated', 'actualCol': 'Investigation Completed on Time', 'kpi_row': 18},
    {'key': 'notification', 'reportId': '1199821531598724', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month', 'plannedCol': 'Total Incident', 'actualCol': 'Notification Submitted on Time', 'kpi_row': 19},
]

# Waste data source
WASTE_SOURCE = {'sheetId': '8150747345538948', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month'}

WASTE_TABLE_COLS = ['General Waste', 'Food Waste', 'Paper Waste', 'Aluminum',
                    'PET Bottle', 'Paper Cup/Carton', 'Single Use Plastic',
                    'Tissue', 'Scrap Metal', 'E-waste', 'Hazardous']
RECYCLABLE_COLS = ['Food Waste', 'Paper Waste', 'Aluminum', 'PET Bottle',
                   'Paper Cup/Carton', 'Single Use Plastic', 'Tissue',
                   'Scrap Metal', 'E-waste']

MONTH_NAMES = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']

# ── Smartsheet API ──

def _ss_fetch(endpoint, token):
    url = f'https://api.smartsheet.com/2.0/{endpoint}'
    req = Request(url, headers={'Authorization': f'Bearer {token}', 'Accept': 'application/json'})
    with urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())

def fetch_sheet_rows(sheet_id, token):
    data = _ss_fetch(f'sheets/{sheet_id}?pageSize=500', token)
    if not data.get('rows'): return []
    col_map = {c['id']: c['title'] for c in data.get('columns', [])}
    rows = []
    for row in data['rows']:
        rec = {}
        for cell in row.get('cells', []):
            title = col_map.get(cell.get('columnId'))
            if title:
                rec[title] = cell.get('displayValue') or cell.get('value') or ''
        if rec:
            rows.append(rec)
    return rows

def fetch_report_rows(report_id, token):
    data = _ss_fetch(f'reports/{report_id}?pageSize=500&level=1', token)
    if not data.get('rows'): return []
    col_map = {}
    for c in data.get('columns', []):
        if c.get('id'): col_map[c['id']] = c['title']
        if c.get('virtualId'): col_map[c['virtualId']] = c['title']
    rows = []
    for row in data['rows']:
        rec = {}
        for cell in row.get('cells', []):
            col_id = cell.get('virtualColumnId') or cell.get('columnId')
            title = col_map.get(col_id)
            if title:
                rec[title] = cell.get('displayValue') or cell.get('value') or ''
        if rec:
            rows.append(rec)
    return rows

def normalize_month(v):
    if not v: return None
    s = str(v).strip()
    if not s: return None
    for m in MONTH_NAMES:
        if m.lower() == s.lower(): return m
    abbr = s[:3].lower()
    abbr_map = {m[:3].lower(): m for m in MONTH_NAMES}
    if abbr in abbr_map: return abbr_map[abbr]
    try:
        from datetime import datetime as dt
        d = dt.strptime(s, '%Y-%m-%d')
        return MONTH_NAMES[d.month - 1]
    except: pass
    try:
        from datetime import datetime as dt
        d = dt.strptime(s, '%m/%d/%Y')
        return MONTH_NAMES[d.month - 1]
    except: pass
    return None

def safe_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def pct_str(v):
    return f"{round(v * 100)}%"


# ── Fetch and process KPI data from Smartsheet ──

def fetch_kpi_data(token, month_filter):
    """Fetch all KPI sources and return {campus_code: {kpi_row: {planned, achieved, calc, weight}}}"""
    data = {}
    for src in SYNC_SOURCES:
        try:
            if src.get('reportId'):
                rows = fetch_report_rows(src['reportId'], token)
            else:
                rows = fetch_sheet_rows(src['sheetId'], token)
        except Exception as e:
            print(f"  WARNING: Failed to fetch {src['key']}: {e}")
            continue

        kpi_row = src['kpi_row']
        campus_col = src['campusCol']
        month_col = src.get('monthCol')
        planned_col = src.get('plannedCol')
        actual_col = src.get('actualCol')
        value_col = src.get('valueCol')
        is_yes_no = src.get('yesNoCount', False)

        # Group by campus, filter by month
        campus_agg = {}  # {campus: {planned: float, actual: float}}
        for row in rows:
            campus = str(row.get(campus_col, '')).strip()
            if not campus: continue

            # Month filter
            if month_col and month_filter:
                row_month = normalize_month(row.get(month_col))
                if not row_month: continue
                # Also try fallback month columns
                if row_month != month_filter:
                    row_month = normalize_month(row.get('Reporting Month'))
                    if row_month != month_filter:
                        row_month = normalize_month(row.get('Date Reported'))
                        if row_month != month_filter:
                            row_month = normalize_month(row.get('Primary'))
                            if row_month != month_filter:
                                continue

            if campus not in campus_agg:
                campus_agg[campus] = {'planned': 0, 'actual': 0}

            if is_yes_no:
                p = 1 if str(row.get(planned_col, '')).strip().lower() == 'yes' else 0
                a = 1 if str(row.get(actual_col, '')).strip().lower() == 'yes' else 0
                campus_agg[campus]['planned'] += p
                campus_agg[campus]['actual'] += a
            elif planned_col and actual_col:
                campus_agg[campus]['planned'] += safe_float(row.get(planned_col))
                campus_agg[campus]['actual'] += safe_float(row.get(actual_col))
            elif value_col:
                v = safe_float(row.get(value_col))
                campus_agg[campus]['planned'] += v
                campus_agg[campus]['actual'] += v

        # Store in data structure
        weight = KPI_WEIGHTS.get(kpi_row, 0.05)
        for campus, agg in campus_agg.items():
            if campus not in data:
                data[campus] = {}
            planned = agg['planned']
            achieved = agg['actual']
            calc = min(achieved / planned, 1.0) if planned > 0 else (1.0 if achieved > 0 else 0)
            data[campus][kpi_row] = {'planned': planned, 'achieved': achieved, 'calc': calc, 'weight': weight}

    return data

def fetch_waste_data(token, month_filter):
    """Fetch waste segregation data from Smartsheet."""
    try:
        rows = fetch_sheet_rows(WASTE_SOURCE['sheetId'], token)
    except:
        return {}
    waste = {}
    for row in rows:
        campus = str(row.get(WASTE_SOURCE['campusCol'], '')).strip()
        month = normalize_month(row.get(WASTE_SOURCE['monthCol']))
        if not campus or month != month_filter: continue
        entry = {}
        for col in WASTE_TABLE_COLS:
            entry[col] = safe_float(row.get(col))
        entry['Total Waste'] = safe_float(row.get('Total Waste'))
        if entry['Total Waste'] == 0:
            entry['Total Waste'] = sum(entry.get(c, 0) for c in WASTE_TABLE_COLS)
        waste[campus] = entry
    return waste


# ── KPI data processing ──

def read_campus_data(kpi_data, sheet_name):
    campus = kpi_data.get(sheet_name, {})
    kpis = []
    pillar_scores = []
    for pillar in PILLAR_KPIS:
        p_kpis = []
        for row in pillar['rows']:
            d = campus.get(row, {'planned': 0, 'achieved': 0, 'calc': 1.0, 'weight': KPI_WEIGHTS.get(row, 0.05)})
            p_kpis.append(d)
            kpis.append(d)
        tw = sum(k['weight'] for k in p_kpis)
        score = sum(k['calc'] * k['weight'] for k in p_kpis) / tw if tw > 0 else 0
        pillar_scores.append(score)
    weights = [p['weight'] for p in PILLAR_KPIS]
    overall = sum(s * w for s, w in zip(pillar_scores, weights))
    return {'sheet': sheet_name, 'kpis': kpis, 'pillar_scores': pillar_scores, 'overall': overall}

def read_region_data(kpi_data, region_cfg):
    campuses = [read_campus_data(kpi_data, s) for s in region_cfg['sheets']]
    n = len(campuses)
    avg_p = [sum(c['pillar_scores'][i] for c in campuses)/n for i in range(5)]
    avg_o = sum(c['overall'] for c in campuses)/n
    return {'campuses': campuses, 'avg_pillar': avg_p, 'avg_overall': avg_o, 'short': region_cfg['short']}


# ── XML chart editing (from kpi_pptx.py) ──

def _set_val_axis_max(xml_str, max_val):
    valax_match = re.search(r'<c:valAx>(.*?)</c:valAx>', xml_str, re.DOTALL)
    if not valax_match: return xml_str
    valax = valax_match.group(0)
    scaling_match = re.search(r'<c:scaling>(.*?)</c:scaling>', valax, re.DOTALL)
    if scaling_match:
        inner = scaling_match.group(1)
        if '<c:max' in inner:
            inner = re.sub(r'<c:max val="[^"]*"/>', f'<c:max val="{max_val}"/>', inner)
        else:
            inner += f'<c:max val="{max_val}"/>'
        inner = re.sub(r'<c:min val="[^"]*"/>', '', inner)
        new_scaling = f'<c:scaling>{inner}</c:scaling>'
        new_valax = valax.replace(scaling_match.group(0), new_scaling)
    else:
        new_valax = valax.replace('<c:valAx>', f'<c:valAx><c:scaling><c:max val="{max_val}"/></c:scaling>')
    return xml_str.replace(valax, new_valax)

def _remove_axis_underline(xml_str):
    xml_str = re.sub(r'<a:u val="[^"]*"/>', '', xml_str)
    xml_str = re.sub(r'<a:u val="[^"]*">[^<]*</a:u>', '', xml_str)
    return xml_str

def update_chart_xml(xml_str, campus_names, achieved_vals, avg_val=None, is_single_series=False, remove_underline=False):
    n_cats = len(campus_names)
    if is_single_series:
        parts = xml_str.split('<c:val>')
        if len(parts) >= 2:
            for pi in range(1, len(parts)):
                val_section = parts[pi].split('</c:val>')[0]
                rest = '</c:val>'.join(parts[pi].split('</c:val>')[1:])
                val_section = re.sub(r'(<c:ptCount val=")\d+(")', f'\\g<1>{n_cats}\\2', val_section)
                existing_pts = list(re.finditer(r'<c:pt idx="\d+">\s*<c:v>[^<]*</c:v>\s*</c:pt>', val_section))
                if existing_pts:
                    val_clean = re.sub(r'<c:pt idx="\d+">\s*<c:v>[^<]*</c:v>\s*</c:pt>', '', val_section)
                    new_pts = ''.join(f'<c:pt idx="{i}"><c:v>{achieved_vals[i]}</c:v></c:pt>' for i in range(n_cats))
                    val_clean = val_clean.replace('</c:numCache>', new_pts + '</c:numCache>')
                    parts[pi] = val_clean + '</c:val>' + rest
                    break
        xml_str = '<c:val>'.join(parts)
        xml_str = _set_val_axis_max(xml_str, 1.0)
        return xml_str

    def _update_ser_values(ser_xml, vals):
        def _replace_num(val_match):
            vx = val_match.group(0)
            vx = re.sub(r'(<c:ptCount val=")\d+(")', f'\\g<1>{n_cats}\\2', vx)
            vx = re.sub(r'<c:pt idx="\d+">\s*<c:v>[^<]*</c:v>\s*</c:pt>', '', vx)
            new_pts = ''.join(f'<c:pt idx="{i}"><c:v>{vals[i]}</c:v></c:pt>' for i in range(n_cats))
            vx = vx.replace('</c:numCache>', new_pts + '</c:numCache>')
            return vx
        return re.sub(r'<c:val>.*?</c:val>', _replace_num, ser_xml, flags=re.DOTALL)

    def _update_str_cache(match_str):
        pt_count_m = re.search(r'<c:ptCount val="(\d+)"/>', match_str)
        if pt_count_m:
            existing_count = int(pt_count_m.group(1))
            if existing_count == 1 and n_cats > 1:
                return match_str
        new_cache = f'<c:ptCount val="{n_cats}"/>'
        for i, name in enumerate(campus_names):
            new_cache += f'<c:pt idx="{i}"><c:v>{name}</c:v></c:pt>'
        return re.sub(r'<c:ptCount val="\d+"/>.*?(?=</c:strCache>)', new_cache, match_str, flags=re.DOTALL)

    xml_str = re.sub(r'<c:strCache>.*?</c:strCache>', lambda m: _update_str_cache(m.group(0)), xml_str, flags=re.DOTALL)

    bar_match = re.search(r'<c:barChart>.*?</c:barChart>', xml_str, re.DOTALL)
    if bar_match:
        bar_xml = bar_match.group(0)
        ser_in_bar = list(re.finditer(r'<c:ser>.*?</c:ser>', bar_xml, re.DOTALL))
        for sm in ser_in_bar:
            old_ser = sm.group(0)
            new_ser = _update_ser_values(old_ser, achieved_vals)
            bar_xml = bar_xml.replace(old_ser, new_ser, 1)
        xml_str = xml_str[:bar_match.start()] + bar_xml + xml_str[bar_match.end():]

    line_match = re.search(r'<c:lineChart>.*?</c:lineChart>', xml_str, re.DOTALL)
    if line_match and avg_val is not None:
        line_xml = line_match.group(0)
        ser_in_line = list(re.finditer(r'<c:ser>.*?</c:ser>', line_xml, re.DOTALL))
        avg_vals = [avg_val] * n_cats
        for sm in ser_in_line:
            old_ser = sm.group(0)
            new_ser = _update_ser_values(old_ser, avg_vals)
            line_xml = line_xml.replace(old_ser, new_ser, 1)
        xml_str = xml_str[:line_match.start()] + line_xml + xml_str[line_match.end():]

    xml_str = _set_val_axis_max(xml_str, 1.0)
    if remove_underline:
        xml_str = _remove_axis_underline(xml_str)
    return xml_str


# ── Slide 6 bar scaling ──

BAR_COLOR_GREEN  = 'C0DD97'
BAR_COLOR_BLUE   = 'B5D4F4'
BAR_COLOR_ORANGE = 'FAC775'
BAR_COLOR_RED    = 'F7C1C1'

def _bar_color(pct):
    p = round(pct * 100)
    if p >= 91: return BAR_COLOR_GREEN
    if p >= 81: return BAR_COLOR_BLUE
    if p >= 71: return BAR_COLOR_ORANGE
    return BAR_COLOR_RED

SLIDE6_BAR_MAP = [
    [(2, 1), (5, 4)],
    [(9, 8), (12, 11)],
    [(16, 15), (19, 18)],
    [(23, 22), (26, 25)],
    [(30, 29), (33, 32)],
]

def update_slide6_bars(xml_str, pct_values):
    shapes = list(re.finditer(r'<p:sp>.*?</p:sp>', xml_str, re.DOTALL))
    if len(shapes) < 36: return xml_str
    replacements = []
    for pi, bar_pairs in enumerate(SLIDE6_BAR_MAP):
        for ci, (fill_idx, bg_idx) in enumerate(bar_pairs):
            if pi >= len(pct_values) or ci >= len(pct_values[pi]): continue
            pct = pct_values[pi][ci]
            bg_shape = shapes[bg_idx].group(0)
            fill_shape = shapes[fill_idx].group(0)
            bg_cx_m = re.search(r'<a:ext cx="(\d+)"', bg_shape)
            if not bg_cx_m: continue
            bg_cx = int(bg_cx_m.group(1))
            new_cx = max(int(bg_cx * max(pct, 0.05)), 1)
            new_fill = re.sub(r'(<a:ext cx=")\d+(")', f'\\g<1>{new_cx}\\2', fill_shape, count=1)
            new_color = _bar_color(pct)
            new_fill = re.sub(r'(<a:srgbClr val=")[A-Fa-f0-9]{6}(")', f'\\g<1>{new_color}\\2', new_fill, count=1)
            replacements.append((shapes[fill_idx].start(), shapes[fill_idx].end(), new_fill))
    for start, end, new_text in sorted(replacements, key=lambda x: x[0], reverse=True):
        xml_str = xml_str[:start] + new_text + xml_str[end:]
    return xml_str


# ── Embedded Excel workbook ──

def create_chart_workbook(categories, bar_values, line_values=None):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws['A1'] = 'Category'
        ws['B1'] = 'Value'
        if line_values is not None:
            ws['C1'] = 'Average'
        for i, cat in enumerate(categories):
            ws.cell(row=i+2, column=1, value=cat)
            ws.cell(row=i+2, column=2, value=bar_values[i])
            if line_values is not None:
                ws.cell(row=i+2, column=3, value=line_values[i] if isinstance(line_values, list) else line_values)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return None

def _rewrite_chart_formulas(xml_str, n_cats, has_line=True):
    last_row = n_cats + 1
    cat_range = f"Sheet1!$A$2:$A${last_row}" if n_cats > 1 else "Sheet1!$A$2"
    bar_range = f"Sheet1!$B$2:$B${last_row}" if n_cats > 1 else "Sheet1!$B$2"
    line_range = f"Sheet1!$C$2:$C${last_row}" if n_cats > 1 else "Sheet1!$C$2"
    ser_idx = [0]
    def replace_ser(ser_match):
        ser_xml = ser_match.group(0)
        idx = ser_idx[0]
        ser_idx[0] += 1
        val_col = 'B' if idx == 0 else 'C'
        ser_xml = re.sub(r'(<c:tx>\s*<c:strRef>\s*<c:f>)[^<]*(</c:f>)', f'\\1Sheet1!${val_col}$1\\2', ser_xml, flags=re.DOTALL)
        ser_xml = re.sub(r'(<c:cat>\s*<c:strRef>\s*<c:f>)[^<]*(</c:f>)', f'\\1{cat_range}\\2', ser_xml, flags=re.DOTALL)
        ser_xml = re.sub(r'(<c:cat>\s*<c:numRef>\s*<c:f>)[^<]*(</c:f>)', f'\\1{cat_range}\\2', ser_xml, flags=re.DOTALL)
        v_range = bar_range if idx == 0 else line_range
        ser_xml = re.sub(r'(<c:val>\s*<c:numRef>\s*<c:f>)[^<]*(</c:f>)', f'\\1{v_range}\\2', ser_xml, flags=re.DOTALL)
        return ser_xml
    def replace_bar(bar_match):
        bar_xml = bar_match.group(0)
        ser_idx[0] = 0
        return re.sub(r'<c:ser>.*?</c:ser>', replace_ser, bar_xml, flags=re.DOTALL)
    xml_str = re.sub(r'<c:barChart>.*?</c:barChart>', replace_bar, xml_str, flags=re.DOTALL)
    def replace_line(line_match):
        line_xml = line_match.group(0)
        return re.sub(r'<c:ser>.*?</c:ser>', replace_ser, line_xml, flags=re.DOTALL)
    if has_line:
        xml_str = re.sub(r'<c:lineChart>.*?</c:lineChart>', replace_line, xml_str, flags=re.DOTALL)
    return xml_str

def update_chart_rels_for_embedding(rels_xml, embed_target):
    return re.sub(
        r'<Relationship\s+Id="rId3"\s+Type="[^"]*oleObject[^"]*"\s+Target="[^"]*"\s+TargetMode="External"\s*/>',
        f'<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/package" Target="{embed_target}"/>',
        rels_xml
    )


# ── Waste slide helpers ──

def _set_cell_text(tc_elem, text, a_ns):
    for r_elem in tc_elem.iter(f'{{{a_ns}}}r'):
        t_elem = r_elem.find(f'{{{a_ns}}}t')
        if t_elem is not None:
            t_elem.text = text
            break
    else:
        for t_elem in tc_elem.iter(f'{{{a_ns}}}t'):
            t_elem.text = text
            break

def fmt_waste(v):
    if v is None or v == 0: return '-'
    if v == int(v): return str(int(v))
    return f'{v:.1f}'.rstrip('0').rstrip('.')

def _blank_waste_slide(file_contents, short_names):
    slide22_path = 'ppt/slides/slide22.xml'
    if slide22_path not in file_contents: return
    xml = file_contents[slide22_path].decode('utf-8')
    tree = ET.ElementTree(ET.fromstring(xml))
    root = tree.getroot()
    a_ns = NS['a']
    for tbl_elem in root.iter(f'{{{a_ns}}}tbl'):
        rows = list(tbl_elem.iter(f'{{{a_ns}}}tr'))
        for ci in range(min(len(short_names), 2)):
            if ci + 2 >= len(rows): break
            data_row = rows[ci + 2]
            cells = list(data_row.iter(f'{{{a_ns}}}tc'))
            _set_cell_text(cells[0], short_names[ci], a_ns)
            for c in cells[1:]:
                _set_cell_text(c, '0', a_ns)
    xml = ET.tostring(root, encoding='unicode')
    xml = re.sub(r'\d+(?:,\d+)*(?:\.\d+)?\s*kg', '0 kg', xml)
    xml = re.sub(r'Recyclable:\s*\d+(?:,\d+)*(?:\.\d+)?\s*kg', 'Recyclable: 0 kg', xml)
    xml = re.sub(r'\d+(?:\.\d+)?%', '0.0%', xml)
    file_contents[slide22_path] = xml.encode('utf-8')

def update_waste_slide(file_contents, region_cfg, short_names, waste_data):
    slide22_path = 'ppt/slides/slide22.xml'
    if slide22_path not in file_contents: return
    campus_sheets = region_cfg['sheets']
    campus_waste = []
    for cs in campus_sheets:
        w = waste_data.get(cs, {})
        entry = {col: w.get(col, 0) for col in WASTE_TABLE_COLS}
        total = w.get('Total Waste', sum(entry.values()))
        recyclable = sum(entry.get(c, 0) for c in RECYCLABLE_COLS)
        recycle_pct = (recyclable / total * 100) if total > 0 else 0.0
        entry['_total'] = total
        entry['_recyclable'] = recyclable
        entry['_recycle_pct'] = recycle_pct
        campus_waste.append(entry)

    xml = file_contents[slide22_path].decode('utf-8')
    tree = ET.ElementTree(ET.fromstring(xml))
    root = tree.getroot()
    a_ns = NS['a']
    col_keys = WASTE_TABLE_COLS

    for tbl_elem in root.iter(f'{{{a_ns}}}tbl'):
        rows_el = list(tbl_elem.iter(f'{{{a_ns}}}tr'))
        for ci in range(min(len(campus_sheets), 2)):
            if ci + 2 >= len(rows_el): break
            w = campus_waste[ci]
            data_row = rows_el[ci + 2]
            cells = list(data_row.iter(f'{{{a_ns}}}tc'))
            _set_cell_text(cells[0], short_names[ci], a_ns)
            for j, key in enumerate(col_keys):
                if j + 1 < len(cells):
                    _set_cell_text(cells[j + 1], fmt_waste(w.get(key, 0)), a_ns)

    xml = ET.tostring(root, encoding='unicode')

    total_replacements = []
    recycle_replacements = []
    pct_replacements = []
    for ci in range(min(len(campus_sheets), 2)):
        w = campus_waste[ci]
        tv = w['_total']
        total_replacements.append(f'{tv:.2f} kg' if tv != int(tv) else f'{int(tv)} kg')
        rv = w['_recyclable']
        recycle_replacements.append(f'Recyclable: {rv:.1f} kg' if rv != int(rv) else f'Recyclable: {int(rv)} kg')
        pct_replacements.append(f'{w["_recycle_pct"]:.1f}%')

    ti = [0]
    def _rt(m):
        if ti[0] < len(total_replacements):
            r = total_replacements[ti[0]]; ti[0] += 1; return r
        return m.group(0)
    xml = re.sub(r'\d+(?:\.\d+)?\s*kg', _rt, xml)

    ri = [0]
    def _rr(m):
        if ri[0] < len(recycle_replacements):
            r = recycle_replacements[ri[0]]; ri[0] += 1; return r
        return m.group(0)
    xml = re.sub(r'Recyclable:\s*\d+(?:\.\d+)?\s*kg', _rr, xml)

    pi = [0]
    def _rp(m):
        if pi[0] < len(pct_replacements):
            r = pct_replacements[pi[0]]; pi[0] += 1; return r
        return m.group(0)
    xml = re.sub(r'\d+(?:\.\d+)?%', _rp, xml)

    file_contents[slide22_path] = xml.encode('utf-8')


# ── Main generation ──

def generate_presentation(template_bytes, region_name, period, kpi_data, waste_data=None):
    if region_name not in REGIONS:
        return None, f"Unknown region: {region_name}"

    region_cfg = REGIONS[region_name]
    region_data = read_region_data(kpi_data, region_cfg)
    campuses = region_data['campuses']
    short_names = region_data['short']

    # Read template ZIP into memory
    buf_in = io.BytesIO(template_bytes)
    file_contents = {}
    with zipfile.ZipFile(buf_in, 'r') as zin:
        for item in zin.infolist():
            file_contents[item.filename] = zin.read(item.filename)

    # 1. Update column charts
    merged_region_name = short_names[0].rsplit(' ', 1)[0] if short_names else region_name
    for chart_file, kpi_idx in CHART_KPI_MAP.items():
        path = f"ppt/charts/{chart_file}"
        if path not in file_contents: continue
        xml_str = file_contents[path].decode('utf-8')
        is_single = chart_file in SINGLE_SERIES_CHARTS
        do_remove_underline = chart_file in REMOVE_UNDERLINE_CHARTS

        if chart_file in MERGE_CAMPUS_CHARTS and len(campuses) > 1:
            avg_calc = sum(c['kpis'][kpi_idx]['calc'] for c in campuses) / len(campuses)
            chart_names = [merged_region_name]
            chart_achieved = [avg_calc]
            chart_avg = avg_calc
        else:
            chart_names = short_names
            chart_achieved = [c['kpis'][kpi_idx]['calc'] for c in campuses]
            chart_avg = sum(c['kpis'][kpi_idx]['calc'] for c in campuses) / len(campuses)

        new_xml = update_chart_xml(xml_str, chart_names, chart_achieved, chart_avg, is_single, remove_underline=do_remove_underline)
        new_xml = _set_val_axis_max(new_xml, 1.0)
        has_line = not is_single
        new_xml = _rewrite_chart_formulas(new_xml, len(chart_names), has_line=has_line)
        file_contents[path] = new_xml.encode('utf-8')

        # Create embedded workbook
        rels_path = f"ppt/charts/_rels/{chart_file}.rels"
        if rels_path in file_contents:
            chart_num = re.search(r'chart(\d+)', chart_file).group(1)
            embed_name = f"Microsoft_Excel_Chart{chart_num}.xlsx"
            embed_path = f"ppt/embeddings/{embed_name}"
            line_vals = None if is_single else [chart_avg] * len(chart_names)
            xlsx_bytes = create_chart_workbook(chart_names, chart_achieved, line_vals)
            if xlsx_bytes:
                file_contents[embed_path] = xlsx_bytes
                rels_xml = file_contents[rels_path].decode('utf-8')
                if 'oleObject' in rels_xml and 'TargetMode="External"' in rels_xml:
                    new_rels = update_chart_rels_for_embedding(rels_xml, f"../embeddings/{embed_name}")
                    file_contents[rels_path] = new_rels.encode('utf-8')

    # 2. Update slides — replace campus name references
    text_replacements = [('Baniyas Campus A', short_names[0] if short_names else ''), ('Baniyas A', short_names[0] if short_names else '')]
    if len(short_names) > 1:
        text_replacements += [('Baniyas Campus B', short_names[1]), ('Baniyas B', short_names[1])]
    else:
        text_replacements += [('Baniyas Campus B', ''), ('Baniyas B', '')]

    for fname in file_contents:
        if fname.startswith('ppt/slides/slide') and fname.endswith('.xml') and fname != 'ppt/slides/slide1.xml':
            xml = file_contents[fname].decode('utf-8')
            changed = False
            for old, new in text_replacements:
                if old in xml:
                    xml = xml.replace(old, new)
                    changed = True
            if 'Q1 Baseline' in xml:
                xml = xml.replace('Q1 Baseline', period)
                changed = True
            if changed:
                file_contents[fname] = xml.encode('utf-8')

    # 3. Update slide 1 (cover)
    slide1_path = 'ppt/slides/slide1.xml'
    if slide1_path in file_contents:
        xml = file_contents[slide1_path].decode('utf-8')
        date_str = datetime.now().strftime('%A, %B %d, %Y')
        xml = re.sub(r'(Wednesday, May 13, 2026|Monday, \w+ \d+, \d{4})', date_str, xml)
        new_sub = region_cfg['subtitle'].replace('&', '&amp;')
        xml = xml.replace('Baniyas Campus A &amp; Campus B', new_sub)
        xml = xml.replace('Baniyas Campus A & Campus B', region_cfg['subtitle'])
        file_contents[slide1_path] = xml.encode('utf-8')

    # 4. Update slide 6 percentage shapes + bars
    slide6_path = 'ppt/slides/slide6.xml'
    if slide6_path in file_contents:
        xml = file_contents[slide6_path].decode('utf-8')
        pct_values = []
        for pi_idx in range(5):
            row_vals = [c['pillar_scores'][pi_idx] for c in campuses]
            row_vals.append(region_data['avg_pillar'][pi_idx])
            pct_values.append(row_vals)
        total_vals = [c['overall'] for c in campuses]
        total_vals.append(region_data['avg_overall'])

        pct_pattern = r'(<a:t>)(\d+%)(</a:t>)'
        matches = list(re.finditer(pct_pattern, xml))
        new_pcts = []
        for pi_idx in range(5):
            for ci in range(min(len(campuses), 2) + 1):
                idx = ci if ci < len(campuses) else len(campuses)
                if idx < len(pct_values[pi_idx]):
                    new_pcts.append(pct_str(pct_values[pi_idx][idx]))
        for ci in range(min(len(campuses), 2) + 1):
            idx = ci if ci < len(campuses) else len(campuses)
            if idx < len(total_vals):
                new_pcts.append(pct_str(total_vals[idx]))

        offset = 0
        pct_idx = 0
        for match in matches:
            if pct_idx >= len(new_pcts): break
            old_val = match.group(2)
            new_val = new_pcts[pct_idx]
            start = match.start(2) + offset
            end = match.end(2) + offset
            xml = xml[:start] + new_val + xml[end:]
            offset += len(new_val) - len(old_val)
            pct_idx += 1

        bar_pcts = []
        for pi_idx in range(5):
            row = [campuses[ci]['pillar_scores'][pi_idx] for ci in range(min(len(campuses), 2))]
            bar_pcts.append(row)
        xml = update_slide6_bars(xml, bar_pcts)
        file_contents[slide6_path] = xml.encode('utf-8')

    # 5. Waste slide
    if waste_data and any(waste_data.values()):
        update_waste_slide(file_contents, region_cfg, short_names, waste_data)
    else:
        _blank_waste_slide(file_contents, short_names)

    # Final pass: force Y-axis max on all charts
    for fname in list(file_contents.keys()):
        if fname.startswith('ppt/charts/chart') and fname.endswith('.xml'):
            xml = file_contents[fname].decode('utf-8') if isinstance(file_contents[fname], bytes) else file_contents[fname]
            valax_m = re.search(r'<c:valAx>(.*?)</c:valAx>', xml, re.DOTALL)
            if valax_m:
                valax = valax_m.group(0)
                sc_m = re.search(r'<c:scaling>(.*?)</c:scaling>', valax, re.DOTALL)
                if sc_m:
                    inner = sc_m.group(1)
                    if '<c:max' in inner:
                        inner = re.sub(r'<c:max val="[^"]*"/>', '<c:max val="1.0"/>', inner)
                    else:
                        inner += '<c:max val="1.0"/>'
                    inner = re.sub(r'<c:min val="[^"]*"/>', '', inner)
                    new_valax = valax.replace(sc_m.group(0), f'<c:scaling>{inner}</c:scaling>')
                    xml = xml.replace(valax, new_valax)
                    file_contents[fname] = xml.encode('utf-8')

    # Write output ZIP
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for fname, data in file_contents.items():
            zout.writestr(fname, data)

    return buf_out.getvalue(), None


# ── HTTP Handler ──

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        params = parse_qs(parsed.query)

        region = params.get('region', ['Abu Dhabi'])[0]
        month = params.get('month', [MONTH_NAMES[datetime.now().month - 2] if datetime.now().month > 1 else 'December'])[0]
        year = params.get('year', [str(datetime.now().year)])[0]

        token = os.environ.get('SMARTSHEET_TOKEN')
        if not token:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'SMARTSHEET_TOKEN not set'}).encode())
            return

        if region not in REGIONS:
            self.send_response(400)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': f'Invalid region. Available: {list(REGIONS.keys())}'}).encode())
            return

        try:
            # Fetch KPI data
            kpi_data = fetch_kpi_data(token, month)

            # Fetch waste data
            waste_data = fetch_waste_data(token, month)

            # Load template
            template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
            template_path = os.path.join(template_dir, 'template.pptx')
            with open(template_path, 'rb') as f:
                template_bytes = f.read()

            period = f"{month} {year}"
            pptx_bytes, error = generate_presentation(template_bytes, region, period, kpi_data, waste_data)

            if error:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': error}).encode())
                return

            safe_name = region.replace(' ', '_')
            filename = f"HCT_KPI_Committee_{safe_name}_{month}_{year}.pptx"

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.presentationml.presentation')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Content-Length', str(len(pptx_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(pptx_bytes)

        except Exception as e:
            import traceback
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e), 'trace': traceback.format_exc()}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
