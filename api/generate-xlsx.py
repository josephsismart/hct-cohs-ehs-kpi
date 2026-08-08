"""Vercel Python serverless function — HCT-COHS KPI Excel Report Generator.
Fetches live data from Smartsheet API and generates downloadable .xlsx files
matching the client template format.
"""

import os, io, json
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from urllib.request import Request, urlopen
from datetime import datetime

MONTH_NAMES = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']

# ── Region mapping (for HS Committee which uses region names) ──
REGIONS = {
    'AD Al Ain':      {'sheets': ['AAF','AAZ']},
    'Abu Dhabi':      {'sheets': ['ADA','ADB']},
    'AD Remote':      {'sheets': ['ADH','MZY']},
    'Dubai':          {'sheets': ['DMC','DBN']},
    'Fujairah':       {'sheets': ['FJF','FJH']},
    'Sharjah':        {'sheets': ['SJA','SJB']},
    'Ras Al Khaimah': {'sheets': ['RKA','RKB']},
}

CAMPUS_TO_REGION = {}
for rname, rcfg in REGIONS.items():
    for code in rcfg['sheets']:
        CAMPUS_TO_REGION[code] = rname

VALID_CAMPUSES = {'AAF','AAZ','ADA','ADB','ADH','MZY','DMC','DBN','FJF','FJH','SJA','SJB','RKA','RKB'}
VALID_REGIONS = set(REGIONS.keys())

# Waste data source
WASTE_SOURCE = {'sheetId': '8150747345538948', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month'}
WASTE_TABLE_COLS = ['General Waste', 'Food Waste', 'Paper Waste', 'Aluminum',
                    'PET Bottle', 'Tissue', 'Scrap Metal', 'E-waste', 'Hazardous']

# ── Smartsheet source definitions ──
# Each entry maps to one sheet in the Excel output
SYNC_SOURCES = [
    {'key': 'v2_hs_committee', 'sheetId': '435993944477572', 'campusCol': 'Committee', 'monthCol': 'Reporting Month',
     'plannedCol': 'Meeting Planned', 'actualCol': 'Meeting Conducted',
     'xlSheet': 'HS Committee Meeting', 'xlType': 'planned_actual', 'useRegion': True},
    {'key': 'v2_external_compliance', 'sheetId': '4198632256393092', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'Applicable Compliance', 'actualCol': 'Actual Compliance',
     'xlSheet': 'External Authority Complianc', 'xlType': 'planned_actual'},
    {'key': 'v2_risk_validated', 'sheetId': '7323092115214212', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'Total Assessments Register', 'actualCol': 'RA Validated and Signed Off',
     'xlSheet': 'Risk Assessment Validated', 'xlType': 'value_actual'},
    {'key': 'v2_findings_on_time', 'sheetId': '4947401822392196', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'No. of Findings in Reporting Month', 'actualCol': 'No. of Findings Due',
     'xlSheet': 'Findings Closed On Time', 'xlType': 'planned_actual'},
    {'key': 'v2_risk_closed', 'sheetId': '7323092115214212', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'Total Risk Assessments Registered', 'actualCol': 'Risk Assessment Closed',
     'xlSheet': 'Risk Assessment Closed', 'xlType': 'planned_actual'},
    {'key': 'v2_ehs_inspection', 'sheetId': '4947401822392196', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'No. of EHS Inspections Planned', 'actualCol': 'No. of EHS Inspections Completed',
     'xlSheet': 'Scheduled EHS Inspection', 'xlType': 'planned_actual'},
    # v2_hs_kpi_report disabled — report 404
    # {'key': 'v2_hs_kpi_report', 'reportId': '4811266391494532', 'campusCol': 'Campuses', 'monthCol': 'Primary',
    #  'valueCol': 'Submitted', 'xlSheet': 'HS KPI Report Submission', 'xlType': 'value'},
    {'key': 'v2_investigation_on_time', 'reportId': '6831846506581892', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'Total Incident Investigated', 'actualCol': 'Investigation Completed on Time',
     'xlSheet': 'Investigation Completed on T', 'xlType': 'planned_actual'},
    {'key': 'notification', 'reportId': '1199821531598724', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'Total Incident', 'actualCol': 'Notification Submitted on Time',
     'xlSheet': 'Incident Notification on Tim', 'xlType': 'planned_actual'},
    {'key': 'training', 'sheetId': '8549734774951812', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'valueCol': 'Total Hours',
     'xlSheet': 'Total Hours of Training', 'xlType': 'value'},
    {'key': 'v2_planned_training', 'sheetId': '8549734774951812', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'Planned (Yes/No)', 'actualCol': 'Planned (Yes/No)',
     'xlSheet': 'Planned Training Report', 'xlType': 'planned_actual', 'yesNoCount': True},
    {'key': 'v2_drills', 'sheetId': '5053158949605252', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'Planned Drill? (Yes/No)', 'actualCol': 'Are there any submission?',
     'xlSheet': 'Drills', 'xlType': 'value_drills', 'yesNoCount': True},
    {'key': 'v2_onsite_induction', 'sheetId': '5899016251330436', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'No. of New Contractors (Individuals)', 'actualCol': 'Contractors Inducted in the Reporting Month',
     'xlSheet': 'Onsite Safety Induction', 'xlType': 'planned_actual'},
    {'key': 'v2_hazard_id', 'sheetId': '7323092115214212', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'Total Controls Identified', 'actualCol': 'Implemented Controls',
     'xlSheet': 'Hazard Identification', 'xlType': 'planned_actual'},
    {'key': 'v2_safe_working', 'sheetId': '1693592581001092', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'No. of SOPs Verified', 'actualCol': 'No. of SOPs Implemented',
     'xlSheet': 'Safe Working Procedure', 'xlType': 'planned_actual'},
    {'key': 'v2_permit_to_work', 'sheetId': '5899016251330436', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'plannedCol': 'No. of PTWs Issued', 'actualCol': 'Total Work Registered',
     'xlSheet': 'Permit to Work', 'xlType': 'planned_actual'},
    {'key': 'v2_mgmt_review_actions', 'sheetId': '3267265049874308', 'campusCol': 'Campus Group', 'monthCol': 'Reporting Month',
     'plannedCol': 'Total Actions', 'actualCol': 'Actions Closed',
     'xlSheet': 'Management Review Actions', 'xlType': 'planned_actual', 'useRegion': True},
]

# Trend sources — these aggregate across all campuses per month
TREND_SOURCES = [
    {'key': 'incidents', 'reportId': '6831846506581892', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'valueCol': 'Total Incident Investigated',
     'xlSheet': 'Trend — Total Incidents', 'mode': 'sum_value'},
    {'key': 'training_trend', 'sheetId': '8549734774951812', 'campusCol': 'Campus Code', 'monthCol': 'Reporting Month',
     'valueCol': 'Total Hours',
     'xlSheet': 'Trend — Training Hours', 'mode': 'sum_value'},
    {'key': 'ehs_trend', 'sheetId': '4947401822392196', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'No. of EHS Inspections Planned', 'actualCol': 'No. of EHS Inspections Completed',
     'xlSheet': 'Trend — EHS Inspection Rate', 'mode': 'pct'},
    {'key': 'compliance_trend', 'sheetId': '4198632256393092', 'campusCol': 'Campus Code', 'monthCol': 'Primary',
     'plannedCol': 'Applicable Compliance', 'actualCol': 'Actual Compliance',
     'xlSheet': 'Trend — Compliance Rate', 'mode': 'pct'},
]

# ── Smartsheet API ──

def _ss_fetch(endpoint, token):
    url = 'https://api.smartsheet.com/2.0/' + endpoint
    req = Request(url, headers={'Authorization': 'Bearer ' + token, 'Accept': 'application/json'})
    with urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())

def fetch_sheet_rows(sheet_id, token):
    data = _ss_fetch('sheets/' + str(sheet_id) + '?pageSize=10000', token)
    if not data.get('rows'):
        return []
    col_map = {}
    for c in data.get('columns', []):
        col_map[c['id']] = c['title']
    rows = []
    for row in data['rows']:
        rec = {}
        for cell in row.get('cells', []):
            title = col_map.get(cell.get('columnId'))
            if title:
                raw = cell.get('value')
                display = cell.get('displayValue')
                # Prefer displayValue for date/text columns, raw for numbers
                if isinstance(raw, (int, float)):
                    rec[title] = raw
                else:
                    rec[title] = display or raw or ''
        if rec:
            rows.append(rec)
    return rows

def fetch_report_rows(report_id, token):
    data = _ss_fetch('reports/' + str(report_id) + '?pageSize=10000&level=1', token)
    if not data.get('rows'):
        return []
    col_map = {}
    for c in data.get('columns', []):
        if c.get('id'):
            col_map[c['id']] = c['title']
        if c.get('virtualId'):
            col_map[c['virtualId']] = c['title']
    rows = []
    for row in data['rows']:
        rec = {}
        for cell in row.get('cells', []):
            col_id = cell.get('virtualColumnId') or cell.get('columnId')
            title = col_map.get(col_id)
            if title:
                raw = cell.get('value')
                display = cell.get('displayValue')
                if isinstance(raw, (int, float)):
                    rec[title] = raw
                else:
                    rec[title] = display or raw or ''
        if rec:
            rows.append(rec)
    return rows

def normalize_month(v):
    if not v:
        return None
    s = str(v).strip()
    if not s:
        return None
    for m in MONTH_NAMES:
        if m.lower() == s.lower():
            return m
    abbr = s[:3].lower()
    abbr_map = {m[:3].lower(): m for m in MONTH_NAMES}
    if abbr in abbr_map:
        return abbr_map[abbr]
    # Try various date formats
    for fmt in ['%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f',
                '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y']:
        try:
            d = datetime.strptime(s.split('+')[0].split('Z')[0], fmt)
            return MONTH_NAMES[d.month - 1]
        except Exception:
            pass
    return None

def safe_float(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except Exception:
        return default

def detect_latest_month():
    return MONTH_NAMES[datetime.now().month - 1]


def process_source(src, rows, month_filter):
    """Aggregate rows by campus, filtered by month. Returns {campus: {planned, actual, value}}"""
    campus_agg = {}
    campus_col = src['campusCol']
    month_col = src.get('monthCol')
    planned_col = src.get('plannedCol')
    actual_col = src.get('actualCol')
    value_col = src.get('valueCol')
    is_yes_no = src.get('yesNoCount', False)
    use_region = src.get('useRegion', False)

    for row in rows:
        campus = str(row.get(campus_col, '')).strip()
        if not campus:
            continue
        raw_cc = str(row.get('Campus Code', '')).strip()
        if campus in ('HQ', 'ADC') or raw_cc in ('HQ', 'ADC'):
            continue

        # Filter by valid campus codes / region names
        if use_region:
            if campus not in VALID_REGIONS:
                continue

        # Month filter
        if month_col and month_filter:
            row_month = normalize_month(row.get(month_col))
            if not row_month:
                row_month = normalize_month(row.get('Reporting Month'))
            if not row_month:
                row_month = normalize_month(row.get('Primary'))
            if row_month != month_filter:
                continue

        # Map campus to region if needed
        if use_region:
            key = campus
        else:
            key = campus

        if key not in campus_agg:
            campus_agg[key] = {'planned': 0, 'actual': 0, 'value': 0}

        if is_yes_no:
            pv = str(row.get(planned_col, '')).strip().lower()
            av = str(row.get(actual_col, '')).strip().lower()
            p = 1 if pv in ('yes', 'true', '1') else 0
            a = 1 if av in ('yes', 'true', '1') else 0
            campus_agg[key]['planned'] += p
            campus_agg[key]['actual'] += a
        else:
            if planned_col:
                campus_agg[key]['planned'] += safe_float(row.get(planned_col))
            if actual_col:
                campus_agg[key]['actual'] += safe_float(row.get(actual_col))
            if value_col:
                campus_agg[key]['value'] += safe_float(row.get(value_col))

    return campus_agg


def process_trend(src, rows):
    """Aggregate rows by month across all campuses. Returns {month_label: value}"""
    month_agg = {}
    campus_col = src['campusCol']
    month_col = src.get('monthCol', 'Reporting Month')
    mode = src['mode']

    for row in rows:
        campus = str(row.get(campus_col, '')).strip()
        if not campus:
            continue
        if campus in ('HQ', 'ADC'):
            continue
        row_month = normalize_month(row.get(month_col))
        if not row_month:
            row_month = normalize_month(row.get('Primary'))
        if not row_month:
            continue

        if row_month not in month_agg:
            month_agg[row_month] = {'planned': 0, 'actual': 0, 'value': 0}

        if mode == 'sum_value':
            month_agg[row_month]['value'] += safe_float(row.get(src.get('valueCol', '')))
        elif mode == 'pct':
            month_agg[row_month]['planned'] += safe_float(row.get(src.get('plannedCol', '')))
            month_agg[row_month]['actual'] += safe_float(row.get(src.get('actualCol', '')))

    return month_agg


def build_xlsx(token, month_filter, year):
    """Build the Excel workbook and return bytes."""
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Cache fetched sheets/reports to avoid duplicate API calls
    _cache = {}
    def cached_fetch(src):
        if src.get('reportId'):
            key = 'r_' + str(src['reportId'])
            if key not in _cache:
                _cache[key] = fetch_report_rows(src['reportId'], token)
            return _cache[key]
        else:
            key = 's_' + str(src['sheetId'])
            if key not in _cache:
                _cache[key] = fetch_sheet_rows(src['sheetId'], token)
            return _cache[key]

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1A1F71')
    header_align = Alignment(horizontal='center', vertical='center')
    data_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    period_label = month_filter + ' ' + str(year) if month_filter else detect_latest_month() + ' ' + str(year)
    if not month_filter:
        month_filter = detect_latest_month()

    # ── Process each KPI source ──
    for src in SYNC_SOURCES:
        print('Processing: ' + src['xlSheet'])
        try:
            rows = cached_fetch(src)
        except Exception as e:
            print('  WARNING: Failed to fetch ' + src['key'] + ': ' + str(e))
            rows = []

        campus_data = process_source(src, rows, month_filter)

        # Create sheet
        ws = wb.create_sheet(title=src['xlSheet'][:31])  # Excel 31-char limit

        xl_type = src['xlType']

        # Write headers based on type
        if xl_type == 'value' or xl_type == 'value_drills':
            headers = ['Campus', 'Value']
        elif xl_type == 'value_actual':
            headers = ['Campus', 'Value']
        else:
            headers = ['Campus', 'Planned', 'Actual']

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Sort campuses, skip those with all-zero values
        campuses = sorted(c for c in campus_data.keys()
                          if any(v != 0 for v in campus_data[c].values()))
        for row_idx, campus in enumerate(campuses, 2):
            d = campus_data[campus]
            ws.cell(row=row_idx, column=1, value=campus).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = Alignment(vertical='center')

            if xl_type == 'value':
                val = d['value'] if d['value'] else d['actual']
                c = ws.cell(row=row_idx, column=2, value=int(val) if val == int(val) else val)
                c.alignment = data_align
                c.border = thin_border
            elif xl_type == 'value_drills':
                val = d['actual']
                c = ws.cell(row=row_idx, column=2, value=int(val) if val == int(val) else val)
                c.alignment = data_align
                c.border = thin_border
            elif xl_type == 'value_actual':
                val = d['actual']
                c = ws.cell(row=row_idx, column=2, value=int(val) if val == int(val) else val)
                c.alignment = data_align
                c.border = thin_border
            else:
                p = d['planned']
                a = d['actual']
                c1 = ws.cell(row=row_idx, column=2, value=int(p) if p == int(p) else p)
                c1.alignment = data_align
                c1.border = thin_border
                c2 = ws.cell(row=row_idx, column=3, value=int(a) if a == int(a) else a)
                c2.alignment = data_align
                c2.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14
        if len(headers) > 2:
            ws.column_dimensions['C'].width = 14

    # ── HS KPI Report Submission (placeholder — report 404) ──
    ws_kpi = wb.create_sheet(title='HS KPI Report Submission')
    for col_idx, h in enumerate(['Campus', 'Submitted'], 1):
        cell = ws_kpi.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws_kpi.column_dimensions['A'].width = 22
    ws_kpi.column_dimensions['B'].width = 14

    # ── Waste Segregation sheet ──
    print('Processing: Waste Segregation')
    try:
        waste_rows = cached_fetch(WASTE_SOURCE)
    except Exception as e:
        print('  WARNING: Failed to fetch waste data: ' + str(e))
        waste_rows = []

    ws_waste = wb.create_sheet(title='Waste Segregation')
    waste_headers = ['Campus'] + WASTE_TABLE_COLS + ['Total Waste']
    for col_idx, h in enumerate(waste_headers, 1):
        cell = ws_waste.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    waste_row_idx = 2
    waste_by_campus = {}
    for row in waste_rows:
        campus = str(row.get(WASTE_SOURCE['campusCol'], '')).strip()
        if not campus or campus not in VALID_CAMPUSES:
            continue
        row_month = normalize_month(row.get(WASTE_SOURCE['monthCol']))
        if row_month != month_filter:
            continue
        entry = {col: safe_float(row.get(col)) for col in WASTE_TABLE_COLS}
        entry['Total Waste'] = safe_float(row.get('Total Waste')) or sum(entry.values())
        waste_by_campus[campus] = entry

    for campus in sorted(waste_by_campus.keys()):
        entry = waste_by_campus[campus]
        ws_waste.cell(row=waste_row_idx, column=1, value=campus).border = thin_border
        for ci, col in enumerate(WASTE_TABLE_COLS, 2):
            val = entry.get(col, 0)
            c = ws_waste.cell(row=waste_row_idx, column=ci, value=round(val, 1) if val else 0)
            c.alignment = data_align
            c.border = thin_border
        total = entry.get('Total Waste', 0)
        c = ws_waste.cell(row=waste_row_idx, column=len(WASTE_TABLE_COLS) + 2, value=round(total, 1) if total else 0)
        c.alignment = data_align
        c.border = thin_border
        waste_row_idx += 1

    for ci in range(1, len(waste_headers) + 1):
        ws_waste.column_dimensions[get_column_letter(ci)].width = 16
    ws_waste.column_dimensions['A'].width = 22

    # ── Trend sheets ──
    for src in TREND_SOURCES:
        print('Processing trend: ' + src['xlSheet'])
        try:
            rows = cached_fetch(src)
        except Exception as e:
            print('  WARNING: Failed to fetch trend ' + src['key'] + ': ' + str(e))
            rows = []

        month_data = process_trend(src, rows)

        ws = wb.create_sheet(title=src['xlSheet'][:31])

        # Headers
        headers = ['Campus', 'Value']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write month rows in chronological order
        ordered_months = [m for m in MONTH_NAMES if m in month_data]
        for row_idx, m in enumerate(ordered_months, 2):
            d = month_data[m]
            label = m + ' ' + str(year)
            ws.cell(row=row_idx, column=1, value=label).border = thin_border

            if src['mode'] == 'sum_value':
                val = d['value']
            else:
                # pct mode: show as rounded percentage
                val = round(d['actual'] / d['planned'] * 100) if d['planned'] > 0 else 0

            c = ws.cell(row=row_idx, column=2, value=val if isinstance(val, int) else round(val, 1))
            c.alignment = data_align
            c.border = thin_border

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        params = parse_qs(parsed.query)

        token = os.environ.get('SMARTSHEET_TOKEN')
        if not token:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'SMARTSHEET_TOKEN not set'}).encode())
            return

        month = params.get('month', [''])[0]
        year = params.get('year', [str(datetime.now().year)])[0]
        report_name = params.get('reportName', ['HCT-COHS KPI Report'])[0]

        if not month:
            month = detect_latest_month()

        try:
            xlsx_bytes = build_xlsx(token, month, year)

            filename = report_name + ' - ' + month + ' ' + year + '.xlsx'
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="' + filename + '"')
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())
